from __future__ import annotations

import html
import math
import re
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.io import ensure_dir, strip_empty, write_csv
from src.utils.korean_names import koreanize_kr_company_name


SECTION_TITLES = [
    ("priority_top", "우선순위_TOP"),
    ("leading_candidates", "선행매매_후보"),
    ("long_term_candidates", "장기투자_후보"),
    ("theme_summary", "테마_요약"),
    ("foreign_flow", "외국인 수급"),
    ("institution_flow_summary", "기관 수급"),
    ("us_52w_highs", "신고가_미국"),
    ("kr_52w_highs", "신고가_한국"),
    ("us_volume_surges", "거래량_급증_미국"),
    ("kr_volume_surges", "거래량_급증_한국"),
    ("famous_13f_changes", "유명기관_13F증감"),
    ("daily_tracking", "일 트래킹"),
]
SECTION_TITLE_BY_KEY = dict(SECTION_TITLES)
LEGACY_SECTION_TITLES = {
    "외국인_수급": SECTION_TITLE_BY_KEY["foreign_flow"],
    "기관수급_요약": SECTION_TITLE_BY_KEY["institution_flow_summary"],
    "일별_트래킹": SECTION_TITLE_BY_KEY["daily_tracking"],
}

STOCK_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("signals", "포착신호"),
    ("close", "종가"),
    ("market_cap", "시가총액"),
    ("volume", "거래량"),
    ("relative_volume", "상대거래량"),
    ("average_volume_30d", "평균거래량"),
    ("high_52w", "52주고가"),
    ("low_52w", "52주저가"),
    ("distance_to_52w_high_pct", "고가괴리"),
    ("position_52w_pct", "52주위치"),
    ("change_pct", "등락률"),
    ("change_from_open_pct", "시가대비"),
    ("gap_pct", "갭"),
    ("beta", "베타"),
    ("sma50_gap_pct", "50일선괴리"),
    ("sma200_gap_pct", "200일선괴리"),
    ("rsi_14", "RSI"),
    ("adx_14", "ADX"),
    ("atr_14", "ATR"),
    ("volatility_d", "일변동성"),
    ("performance_1w", "1주성과"),
    ("performance_1m", "1개월성과"),
    ("performance_3m", "3개월성과"),
    ("performance_6m", "6개월성과"),
    ("performance_ytd", "YTD성과"),
    ("performance_1y", "1년성과"),
    ("investment_priority_score", "투자우선점수"),
    ("long_future_score", "장기/미래점수"),
    ("leading_supply_score", "선행수급점수"),
    ("foreign_flow_investment_score", "외국인수급점수"),
    ("future_industry_theme", "미래산업테마"),
    ("core_basis", "핵심근거"),
    ("forward_per", "Forward PER"),
    ("forward_peg", "Forward PEG"),
    ("trailing_per", "PER"),
    ("pbr", "PBR"),
    ("price_to_sales", "PSR"),
    ("price_to_fcf", "P/FCF"),
    ("dividend_yield", "배당수익률"),
    ("eps_ttm", "EPS TTM"),
    ("forward_eps", "Forward EPS"),
    ("total_revenue", "매출액"),
    ("revenue_growth_yoy", "매출성장률 YoY"),
    ("revenue_growth_qoq", "매출성장률 QoQ"),
    ("eps_growth_yoy", "EPS성장률 YoY"),
    ("eps_growth_qoq", "EPS성장률 QoQ"),
    ("expected_revenue_growth", "예상매출성장률"),
    ("expected_eps_growth", "예상EPS성장률"),
    ("gross_margin", "매출총이익률"),
    ("operating_margin", "영업이익률"),
    ("profit_margin", "순이익률"),
    ("fcf_margin", "FCF마진"),
    ("free_cash_flow", "FCF"),
    ("roic", "ROIC"),
    ("roe", "ROE"),
    ("roa", "ROA"),
    ("debt_to_equity", "부채비율"),
    ("current_ratio", "유동비율"),
    ("quick_ratio", "당좌비율"),
    ("float_shares", "유통주식수"),
    ("shares_outstanding", "발행주식수"),
    ("employees", "직원수"),
    ("institutional_ownership_pct", "기관보유율"),
    ("insider_ownership_pct", "내부자보유율"),
    ("short_percent_float", "공매도비율"),
    ("short_ratio", "Short Ratio"),
    ("target_upside_pct", "목표가상승여력"),
    ("analyst_opinion", "투자의견"),
    ("recent_report_broker", "최근리포트증권사"),
    ("recent_report_title", "최근리포트제목"),
    ("supply_pattern", "수급패턴"),
    ("foreign_net_buy_5d", "외국인 5일"),
    ("foreign_net_buy_20d", "외국인 20일"),
    ("foreign_net_buy", "외국인순매수"),
    ("institution_net_buy_5d", "기관 5일"),
    ("institution_net_buy_20d", "기관 20일"),
    ("institution_net_buy", "기관순매수"),
    ("net_supply_5d", "합산수급 5일"),
    ("net_supply_20d", "합산수급 20일"),
    ("foreign_ownership_rate", "외국인지분율"),
    ("foreign_ownership_change_20d", "외국인지분율 변화"),
    ("exchange", "거래소"),
    ("sector", "섹터"),
    ("industry", "산업"),
]

PRIORITY_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("market_cap", "시가총액"),
    ("investment_priority_score", "투자우선점수"),
    ("future_industry_theme", "미래산업테마"),
    ("core_basis", "핵심근거"),
    ("forward_per", "Forward PER"),
    ("forward_peg", "Forward PEG"),
    ("price_to_sales", "PSR"),
    ("price_to_fcf", "P/FCF"),
    ("revenue_growth_yoy", "매출성장률 YoY"),
    ("eps_growth_yoy", "EPS성장률 YoY"),
    ("fcf_margin", "FCF마진"),
    ("roic", "ROIC"),
    ("target_upside_pct", "목표가상승여력"),
    ("analyst_opinion", "투자의견"),
    ("recent_report_broker", "최근리포트증권사"),
    ("recent_report_title", "최근리포트제목"),
]

LEADING_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("signals", "포착신호"),
    ("close", "종가"),
    ("leading_supply_score", "선행수급점수"),
    ("relative_volume", "상대거래량"),
    ("change_pct", "등락률"),
    ("gap_pct", "갭"),
    ("performance_1w", "1주성과"),
    ("performance_1m", "1개월성과"),
    ("rsi_14", "RSI"),
    ("adx_14", "ADX"),
    ("foreign_net_buy_5d", "외국인 5일"),
    ("institution_net_buy_5d", "기관 5일"),
    ("net_supply_5d", "합산수급 5일"),
    ("supply_pattern", "수급패턴"),
    ("core_basis", "핵심근거"),
]

LONG_TERM_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("market_cap", "시가총액"),
    ("long_future_score", "장기/미래점수"),
    ("future_industry_theme", "미래산업테마"),
    ("total_revenue", "매출액"),
    ("revenue_growth_yoy", "매출성장률 YoY"),
    ("expected_revenue_growth", "예상매출성장률"),
    ("eps_growth_yoy", "EPS성장률 YoY"),
    ("expected_eps_growth", "예상EPS성장률"),
    ("gross_margin", "매출총이익률"),
    ("operating_margin", "영업이익률"),
    ("profit_margin", "순이익률"),
    ("fcf_margin", "FCF마진"),
    ("free_cash_flow", "FCF"),
    ("roe", "ROE"),
    ("roa", "ROA"),
    ("roic", "ROIC"),
    ("debt_to_equity", "부채비율"),
    ("current_ratio", "유동비율"),
    ("quick_ratio", "당좌비율"),
    ("beta", "베타"),
    ("dividend_yield", "배당수익률"),
]

FOREIGN_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("market_cap", "시가총액"),
    ("foreign_flow_investment_score", "외국인수급점수"),
    ("foreign_net_buy", "외국인순매수"),
    ("foreign_net_buy_amount_mil_krw", "외국인금액(백만)"),
    ("foreign_net_buy_5d", "외국인 5일"),
    ("foreign_net_buy_20d", "외국인 20일"),
    ("foreign_ownership_rate", "외국인지분율"),
    ("foreign_ownership_change_20d", "지분율 20일 변화"),
    ("trailing_per", "PER"),
    ("forward_per", "Forward PER"),
    ("forward_peg", "Forward PEG"),
    ("pbr", "PBR"),
    ("dividend_yield", "배당수익률"),
    ("roic", "ROIC"),
    ("roe", "ROE"),
    ("volume", "거래량"),
    ("foreign_rank_volume", "랭킹거래량"),
    ("relative_volume", "상대거래량"),
    ("supply_source", "수급출처"),
    ("supply_pattern", "수급패턴"),
    ("recent_report_title", "최근리포트제목"),
]

INSTITUTION_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("market_cap", "시가총액"),
    ("institution_flow_score", "기관수급점수"),
    ("institution_net_buy", "기관순매수"),
    ("institution_net_buy_amount_mil_krw", "기관금액(백만)"),
    ("institution_net_buy_5d", "기관 5일"),
    ("institution_net_buy_20d", "기관 20일"),
    ("net_supply_5d", "합산수급 5일"),
    ("net_supply_20d", "합산수급 20일"),
    ("trailing_per", "PER"),
    ("forward_per", "Forward PER"),
    ("forward_peg", "Forward PEG"),
    ("pbr", "PBR"),
    ("dividend_yield", "배당수익률"),
    ("roic", "ROIC"),
    ("roe", "ROE"),
    ("volume", "거래량"),
    ("institution_rank_volume", "랭킹거래량"),
    ("relative_volume", "상대거래량"),
    ("supply_source", "수급출처"),
    ("supply_pattern", "수급패턴"),
    ("recent_report_title", "최근리포트제목"),
]

HIGH_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("high_52w", "52주고가"),
    ("distance_to_52w_high_pct", "고가괴리"),
    ("position_52w_pct", "52주위치"),
    ("market_cap", "시가총액"),
    ("beta", "베타"),
    ("sma50_gap_pct", "50일선괴리"),
    ("sma200_gap_pct", "200일선괴리"),
    ("performance_1m", "1개월성과"),
    ("performance_3m", "3개월성과"),
    ("performance_ytd", "YTD성과"),
    ("volume", "거래량"),
    ("relative_volume", "상대거래량"),
    ("core_basis", "핵심근거"),
]

VOLUME_COLUMNS = [
    ("date", "날짜"),
    ("country", "국가"),
    ("ticker", "티커"),
    ("company_name", "기업명"),
    ("close", "종가"),
    ("volume", "거래량"),
    ("average_volume_30d", "평균거래량"),
    ("relative_volume", "상대거래량"),
    ("change_pct", "등락률"),
    ("change_from_open_pct", "시가대비"),
    ("gap_pct", "갭"),
    ("rsi_14", "RSI"),
    ("adx_14", "ADX"),
    ("atr_14", "ATR"),
    ("volatility_d", "일변동성"),
    ("performance_1w", "1주성과"),
    ("performance_1m", "1개월성과"),
    ("short_percent_float", "공매도비율"),
    ("float_shares", "유통주식수"),
    ("core_basis", "핵심근거"),
]

THEME_COLUMNS = [
    ("date", "날짜"),
    ("future_industry_theme", "테마"),
    ("stock_count", "포착종목수"),
    ("avg_investment_priority_score", "평균 투자우선점수"),
    ("avg_long_future_score", "평균 장기/미래점수"),
    ("avg_leading_supply_score", "평균 선행수급점수"),
    ("avg_relative_volume", "평균 상대거래량"),
    ("top_tickers", "대표 티커"),
    ("core_basis", "포착 종목"),
]

F13_COLUMNS = [
    ("date", "날짜"),
    ("company_name", "보유종목/발행사"),
    ("ticker", "티커"),
    ("new_institution_count", "신규기관수"),
    ("increased_institution_count", "증가기관수"),
    ("position_weight_increased_count", "비중증가기관수"),
    ("max_position_weight_change_pct", "최대비중증가"),
    ("average_position_weight_change_pct", "평균비중증감"),
    ("guru_position_score", "대가비중점수"),
    ("top_position_weight_changes", "비중증가대가"),
    ("decreased_institution_count", "감소기관수"),
    ("exited_institution_count", "청산기관수"),
    ("total_share_change", "총증감주식"),
    ("average_change_pct", "평균증감률"),
    ("total_current_shares", "총현재보유주식"),
    ("famous_13f_score", "13F투자후보점수"),
    ("institutions", "기관"),
    ("core_basis", "보유증감근거"),
]

TRACKING_COLUMNS = [
    ("date", "날짜"),
    ("section", "섹션"),
    ("row_count", "행수"),
    ("generated_at_kst", "생성시각"),
]

SECTION_COLUMNS = {
    "priority_top": PRIORITY_COLUMNS,
    "leading_candidates": LEADING_COLUMNS,
    "long_term_candidates": LONG_TERM_COLUMNS,
    "theme_summary": THEME_COLUMNS,
    "foreign_flow": FOREIGN_COLUMNS,
    "institution_flow_summary": INSTITUTION_COLUMNS,
    "us_52w_highs": HIGH_COLUMNS,
    "kr_52w_highs": HIGH_COLUMNS,
    "us_volume_surges": VOLUME_COLUMNS,
    "kr_volume_surges": VOLUME_COLUMNS,
    "famous_13f_changes": F13_COLUMNS,
    "daily_tracking": TRACKING_COLUMNS,
}

SUMMARY_FIELDS = ["date", "section", "row_count", "generated_at_kst"]
NUMERIC_FIELDS = {
    "close",
    "market_cap",
    "volume",
    "average_volume_30d",
    "high_52w",
    "low_52w",
    "investment_priority_score",
    "long_future_score",
    "leading_supply_score",
    "institution_flow_score",
    "foreign_flow_investment_score",
    "stock_count",
    "avg_investment_priority_score",
    "avg_long_future_score",
    "avg_leading_supply_score",
    "avg_relative_volume",
    "beta",
    "rsi_14",
    "adx_14",
    "atr_14",
    "forward_per",
    "forward_peg",
    "trailing_per",
    "pbr",
    "price_to_sales",
    "price_to_fcf",
    "eps_ttm",
    "forward_eps",
    "total_revenue",
    "free_cash_flow",
    "debt_to_equity",
    "current_ratio",
    "quick_ratio",
    "float_shares",
    "shares_outstanding",
    "employees",
    "short_ratio",
    "foreign_net_buy_5d",
    "foreign_net_buy_20d",
    "foreign_net_buy",
    "foreign_net_buy_amount_mil_krw",
    "foreign_rank_volume",
    "institution_net_buy_5d",
    "institution_net_buy_20d",
    "institution_net_buy",
    "institution_net_buy_amount_mil_krw",
    "institution_rank_volume",
    "net_supply_5d",
    "net_supply_20d",
    "new_institution_count",
    "increased_institution_count",
    "position_weight_increased_count",
    "position_weight_decreased_count",
    "guru_position_score",
    "decreased_institution_count",
    "exited_institution_count",
    "total_share_change",
    "total_current_shares",
    "famous_13f_score",
    "row_count",
}
PERCENT_FIELDS = {
    "change_pct",
    "change_from_open_pct",
    "gap_pct",
    "distance_to_52w_high_pct",
    "position_52w_pct",
    "sma50_gap_pct",
    "sma200_gap_pct",
    "volatility_d",
    "performance_1w",
    "performance_1m",
    "performance_3m",
    "performance_6m",
    "performance_ytd",
    "performance_1y",
    "dividend_yield",
    "revenue_growth_yoy",
    "revenue_growth_qoq",
    "eps_growth_yoy",
    "eps_growth_qoq",
    "expected_revenue_growth",
    "expected_eps_growth",
    "gross_margin",
    "operating_margin",
    "profit_margin",
    "fcf_margin",
    "roic",
    "roe",
    "roa",
    "institutional_ownership_pct",
    "insider_ownership_pct",
    "short_percent_float",
    "target_upside_pct",
    "foreign_ownership_rate",
    "foreign_ownership_change_20d",
    "average_change_pct",
    "max_position_weight_change_pct",
    "average_position_weight_change_pct",
    "current_position_weight_pct",
    "previous_position_weight_pct",
    "position_weight_change_pct",
}
HIDDEN_DISPLAY_FIELDS = {"report_link", "source_url"}


def write_outputs(
    root: Path,
    run_date: date,
    sections: dict[str, list[dict[str, Any]]],
    records: list[dict[str, Any]],
    generated_at_kst: str,
) -> None:
    processed_dir = ensure_dir(root / "data" / "processed")
    reports_dir = ensure_dir(root / "reports")
    dated_report_dir = ensure_dir(reports_dir / run_date.isoformat())

    sections = {key: _dedupe_rows(value) for key, value in sections.items()}
    summary_rows = [
        {
            "date": run_date.isoformat(),
            "section": title,
            "row_count": len(sections.get(key, [])),
            "generated_at_kst": generated_at_kst,
        }
        for key, title in SECTION_TITLES
    ]
    sections["daily_tracking"] = _merge_daily_tracking(processed_dir / "daily_summary.csv", summary_rows)

    payload = strip_empty(
        {
            "generated_at_kst": generated_at_kst,
            "date": run_date.isoformat(),
            "sections": sections,
        }
    )

    write_csv(processed_dir / "daily_summary.csv", sections["daily_tracking"], SUMMARY_FIELDS)
    write_csv(reports_dir / "daily_summary.csv", sections["daily_tracking"], SUMMARY_FIELDS)
    _write_xlsx(reports_dir / "latest.xlsx", sections)

    html_text = render_html(payload)
    (root / "index.html").write_text(html_text, encoding="utf-8")
    (reports_dir / "report.html").write_text(html_text, encoding="utf-8")
    (dated_report_dir / "report.html").write_text(html_text, encoding="utf-8")


def render_html(payload: dict[str, Any]) -> str:
    generated = payload.get("generated_at_kst") or ""
    generated_label = _format_generated_label(str(generated))
    sections = payload.get("sections") or {}
    tab_buttons = "\n".join(
        f"<button class='tab-btn{' on' if index == 0 else ''}' type='button' data-tab='{key}'>{html.escape(_tab_title(key, title, generated_label))}</button>"
        for index, (key, title) in enumerate(SECTION_TITLES)
    )
    panels = "\n".join(
        _render_panel(key, title, sections.get(key, []), active=(index == 0))
        for index, (key, title) in enumerate(SECTION_TITLES)
    )
    return f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Stock Trend Report</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;800&family=JetBrains+Mono:wght@500;700&display=swap" rel="stylesheet">
<style>
@font-face{{font-family:'SOK500';src:local('SamsungOneKorean500'),local('SamsungOneKorean 500');font-weight:500;font-style:normal;font-display:swap}}
:root{{--bg:#F5F7FA;--card:#FFFFFF;--card2:#F0FDF4;--hdr:#1A3A2A;--hdrText:#FFFFFF;--row2:#F2F6FC;--t1:#111827;--t2:#475569;--t3:#64748B;--bd:#BBF7D0;--grid:#DDE7F3;--ac:#16A34A;--acL:#DCFCE7;--acT:#15803D;--hoverBg:#BBF7D0;--hoverFg:#052E16;--okB:#15803D;--okT:#FFFFFF;--buyB:#DCFCE7;--buyT:#166534;--warnB:#FEE2E2;--warnT:#991B1B;--negB:#FCA5A5;--negT:#7F1D1D;--link:#0563C1;--tbg:rgba(255,255,255,.92);--r:12px;--shadow:0 4px 18px rgba(15,23,42,.08);--scroll:#86EFAC;--scrollTrack:#ECFDF5}}
[data-t=dark]{{--bg:#071510;--card:#0D1F15;--card2:#122B1C;--row2:#0F1824;--t1:#E2E8F0;--t2:#CBD5E1;--t3:#94A3B8;--bd:#1A3A2A;--grid:#243044;--hdr:#122B1C;--tbg:rgba(7,21,16,.92);--hoverBg:#166534;--hoverFg:#ECFDF5;--scroll:#4ADE80;--scrollTrack:#0B1C13;--shadow:0 8px 26px rgba(0,0,0,.35)}}
*,*::before,*::after{{box-sizing:border-box}}html{{scroll-behavior:smooth;font-size:14px}}body{{margin:0;font-family:'Noto Sans KR','SOK500','Malgun Gothic',system-ui,sans-serif;background:var(--bg);color:var(--t1);line-height:1.55;font-size:1rem;-webkit-font-smoothing:antialiased;text-rendering:optimizeLegibility}}button,input{{font:inherit}}a{{color:var(--link);font-weight:800;text-decoration:none}}a:hover{{text-decoration:underline}}
.topbar{{position:sticky;top:0;z-index:50;background:var(--tbg);backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);border-bottom:1px solid var(--bd);padding:.35rem .9rem;display:flex;align-items:center;gap:.55rem;overflow-x:auto;scrollbar-width:thin;scrollbar-color:var(--scroll) transparent}}
.tabs{{display:flex;gap:.2rem;overflow-x:auto;flex:1;min-width:0;scrollbar-width:thin;scrollbar-color:var(--scroll) transparent}}.tab-btn{{border:1px solid var(--bd);background:var(--card);color:var(--t2);border-radius:999px;padding:.28rem .65rem;font-size:.68rem;font-weight:900;white-space:nowrap;cursor:pointer;transition:background .18s,border-color .18s,color .18s,box-shadow .18s}}.tab-btn:hover{{background:var(--acL);border-color:var(--ac);color:var(--acT)}}.tab-btn.on{{background:var(--ac);border-color:var(--ac);color:#fff;box-shadow:0 2px 10px rgba(22,163,74,.20)}}
.tools{{display:flex;align-items:center;gap:.35rem;flex-shrink:0}}.mode{{border:1px solid var(--bd);background:var(--card2);border-radius:999px;padding:.18rem .52rem;font-size:.66rem;font-weight:800;color:var(--t2);cursor:pointer}}
main{{width:min(1760px,calc(100% - 1.5rem));margin:0 auto;padding:.75rem 0 2rem}}.panel{{display:none;background:var(--card);border:1px solid var(--bd);border-radius:var(--r);box-shadow:var(--shadow);overflow:hidden}}.panel.on{{display:block}}.panel-head{{display:flex;align-items:flex-start;justify-content:space-between;gap:1rem;padding:.62rem 1rem;border-bottom:1px solid var(--bd);background:var(--card2)}}.panel-head h2{{font-size:.9rem;margin:0;font-weight:900}}.panel-head p{{margin:.05rem 0 0;font-size:.7rem;color:var(--t3);font-weight:800}}.downloads{{display:flex;flex-wrap:wrap;gap:.35rem}}.downloads a{{font-size:.68rem;border:1px solid var(--bd);border-radius:6px;padding:.18rem .48rem;background:var(--card);color:var(--acT)}}
.table-wrap{{overflow:auto;max-height:calc(100vh - 132px);scrollbar-gutter:stable both-edges;scrollbar-width:thin;scrollbar-color:var(--scroll) var(--scrollTrack)}}.table-wrap::-webkit-scrollbar,.tabs::-webkit-scrollbar,.topbar::-webkit-scrollbar{{width:8px;height:8px}}.table-wrap::-webkit-scrollbar-track,.tabs::-webkit-scrollbar-track,.topbar::-webkit-scrollbar-track{{background:var(--scrollTrack)}}.table-wrap::-webkit-scrollbar-thumb,.tabs::-webkit-scrollbar-thumb,.topbar::-webkit-scrollbar-thumb{{background:var(--scroll);border-radius:999px;border:2px solid var(--scrollTrack)}}.table-wrap::-webkit-scrollbar-thumb:hover,.tabs::-webkit-scrollbar-thumb:hover,.topbar::-webkit-scrollbar-thumb:hover{{background:var(--ac)}}
table{{width:max-content;min-width:100%;border-collapse:separate;border-spacing:0;font-size:.72rem;line-height:1.34}}th{{position:sticky;top:0;z-index:3;background:var(--hdr);color:var(--hdrText);padding:.34rem .42rem;text-align:left;font-weight:900;white-space:nowrap;border-bottom:2px solid rgba(255,255,255,.22);border-right:1px solid rgba(255,255,255,.14)}}td{{padding:.24rem .42rem;border-bottom:1px solid var(--grid);border-right:1px solid var(--grid);vertical-align:top;background:var(--card);color:var(--t1)}}tbody tr:nth-child(even) td{{background:var(--row2)}}tbody tr:hover td{{background:var(--hoverBg)!important;color:var(--hoverFg)!important;box-shadow:inset 0 1px 0 var(--ac),inset 0 -1px 0 var(--ac)}}tbody tr:hover td:first-child{{box-shadow:inset 4px 0 0 var(--ac),inset 0 1px 0 var(--ac),inset 0 -1px 0 var(--ac)}}tbody tr:hover a{{color:#034EA2!important;text-decoration:underline}}[data-t=dark] tbody tr:hover a{{color:#BFDBFE!important}}.empty{{padding:1.2rem;color:var(--t3);font-size:.82rem;font-weight:800}}.num{{font-family:'JetBrains Mono',monospace;white-space:nowrap;font-weight:650}}.score-cell,.valuation-cell,.growth-cell,.quality-cell,.momentum-cell,.liquidity-cell,.size-cell,.ownership-cell{{background:#F8FAFC!important;color:#334155!important;font-weight:700}}.risk-cell{{background:#FEF2F2!important;color:#991B1B!important;font-weight:750}}.pos-strong{{background:#15803D!important;color:#FFFFFF!important;font-weight:900;border-radius:4px}}.pos-buy{{background:#DCFCE7!important;color:#166534!important;font-weight:850;border-radius:4px}}.warn{{background:#FEE2E2!important;color:#991B1B!important;font-weight:900;border-radius:4px}}.neg{{background:#FCA5A5!important;color:#7F1D1D!important;font-weight:900;border-radius:4px}}.basis{{min-width:260px;max-width:430px}}.bullets{{display:grid;gap:.08rem;line-height:1.32;font-size:.71rem}}.basis-cell .bullets{{max-height:3.9rem;overflow:hidden}}.bullets span{{display:block;padding:.05rem .2rem .05rem .72rem;position:relative;border-left:2px solid transparent;border-radius:4px;font-weight:500}}.bullets span::before{{content:'-';position:absolute;left:.2rem;color:var(--ac);font-weight:800}}.bullets .b-hot{{background:#DCFCE7;border-left-color:#15803D}}.bullets .b-flow{{background:#ECFDF5;border-left-color:#16A34A}}.bullets .b-good{{background:#DCFCE7;border-left-color:#15803D}}.bullets .b-tech{{background:#F0FDF4;border-left-color:#22C55E}}.bullets .b-risk{{background:#FEE2E2;border-left-color:#DC2626}}.kw{{font-weight:900}}.kw-hot{{color:#15803D}}.kw-flow{{color:#047857}}.kw-good{{color:#166534}}.kw-tech{{color:#16A34A}}.kw-risk{{color:#DC2626}}.tag{{display:inline-block;border:1px solid var(--bd);background:var(--card2);border-radius:999px;padding:.04rem .34rem;margin:.04rem;font-size:.64rem;font-weight:800;color:var(--t2)}}footer{{font-size:.66rem;color:var(--t3);text-align:center;padding:.8rem 0}}
.key-cell{{background:#F0FDF4!important;color:#14532D!important;font-weight:900}}.company-cell{{background:#FFFFFF!important;color:#111827!important;font-weight:700;font-size:.73rem;max-width:150px;min-width:92px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.country-us{{background:#F8FAFC!important;color:#334155!important;font-weight:850}}.country-kr{{background:#ECFDF5!important;color:#166534!important;font-weight:850}}.signal-cell{{background:#F8FAFC!important;color:#334155!important;max-width:170px;font-weight:700}}.theme-cell{{background:#F0FDF4!important;color:#166534!important;font-weight:800;max-width:150px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.basis-cell{{background:#FAFAFA!important;color:#334155!important;font-weight:500}}.report-cell{{background:#EFF6FF!important;color:#0563C1!important;font-weight:750;max-width:280px;min-width:150px}}.report-cell a{{display:block;max-width:260px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.volume-hot{{background:#15803D!important;color:#FFFFFF!important;font-weight:900}}.volume-watch{{background:#DCFCE7!important;color:#166534!important;font-weight:850}}.flow-pos{{background:#16A34A!important;color:#FFFFFF!important;font-weight:900}}.metric-blue,.metric-cyan,.metric-purple{{background:#DCFCE7!important;color:#166534!important;font-weight:850}}.metric-soft{{background:#F0FDF4!important;color:#166534!important;font-weight:700}}.section-cell{{background:#ECFDF5!important;color:#047857!important;font-weight:850}}
@media(max-width:768px){{main{{width:calc(100% - .5rem)}}.topbar{{padding:.28rem .45rem;align-items:flex-start;flex-direction:column}}.tools{{align-self:flex-end}}.table-wrap{{max-height:none}}}}
</style>
</head>
<body data-t="light">
<div class="topbar">
  <div class="tabs" role="tablist" aria-label="리포트 섹션">
    {tab_buttons}
  </div>
  <div class="tools">
    <button class="mode" type="button" id="modeToggle">Dark</button>
  </div>
</div>
<main>
  {panels}
</main>
<script>
const tabs=[...document.querySelectorAll('.tab-btn')];
const panels=[...document.querySelectorAll('.panel')];
tabs.forEach(btn=>btn.addEventListener('click',()=>{{
  tabs.forEach(item=>item.classList.toggle('on',item===btn));
  panels.forEach(panel=>panel.classList.toggle('on',panel.id==='panel-'+btn.dataset.tab));
  window.scrollTo({{top:0,behavior:'smooth'}});
}}));
document.getElementById('modeToggle').addEventListener('click',()=>{{
  const next=document.body.getAttribute('data-t')==='dark'?'light':'dark';
  document.body.setAttribute('data-t',next);
  document.getElementById('modeToggle').textContent=next==='dark'?'Light':'Dark';
}});
</script>
</body>
</html>"""


def _tab_title(key: str, title: str, generated_label: str) -> str:
    if key == "daily_tracking" and generated_label:
        return f"{title} {generated_label}"
    return title


def _format_generated_label(value: str) -> str:
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{2}:\d{2})(?::\d{2})?", value.strip())
    if not match:
        return ""
    month = int(match.group(2))
    day = int(match.group(3))
    return f"{month}/{day} {match.group(4)} 기준"


def _render_panel(key: str, title: str, rows: list[dict[str, Any]], active: bool) -> str:
    rows = _dedupe_rows(rows)
    table = _render_table(key, rows)
    return f"""<section class="panel{' on' if active else ''}" id="panel-{html.escape(key)}">
  <div class="panel-head">
    <div><p>{len(rows)}개 유니크 항목</p></div>
    <div class="downloads"><a href="reports/latest.xlsx">latest.xlsx</a></div>
  </div>
  {table}
</section>"""


def _render_table(section_key: str, rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "<div class='empty'>표시할 데이터가 없습니다.</div>"
    columns = _columns_for_rows(SECTION_COLUMNS.get(section_key, STOCK_COLUMNS), rows)
    head = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body = "\n".join(_render_row(row, columns) for row in rows)
    return f"<div class='table-wrap'><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"


def _display_columns(columns: list[tuple[str, str]]) -> list[tuple[str, str]]:
    return [(field, label) for field, label in columns if field not in HIDDEN_DISPLAY_FIELDS]


def _columns_for_rows(columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> list[tuple[str, str]]:
    visible = []
    for field, label in _display_columns(columns):
        if _field_has_display(field, rows):
            visible.append((field, label))
    return visible


def _field_has_display(field: str, rows: list[dict[str, Any]]) -> bool:
    always = {
        "date",
        "country",
        "ticker",
        "company_name",
        "future_industry_theme",
        "section",
        "row_count",
        "generated_at_kst",
    }
    if field in always:
        return True
    for row in rows:
        if row.get(field) not in (None, "", []):
            return True
        if field == "recent_report_broker" and _fallback_report_broker(row):
            return True
        if field == "recent_report_title" and _fallback_report_title(row) and _fallback_report_link(row):
            return True
    return False


def _render_row(row: dict[str, Any], columns: list[tuple[str, str]]) -> str:
    cells = []
    for field, _ in columns:
        value = row.get(field)
        classes = _classes(field, value)
        title = _cell_title(field, value, row)
        title_attr = f" title='{html.escape(title, quote=True)}'" if title else ""
        cells.append(f"<td class='{classes}'{title_attr}>{_format_cell(field, value, row)}</td>")
    return "<tr>" + "".join(cells) + "</tr>"


def _format_cell(field: str, value: Any, row: dict[str, Any]) -> str:
    if field == "recent_report_broker":
        return html.escape(_display_report_broker(row) or "")
    if field == "recent_report_title":
        link = _fallback_report_link(row)
        title = _display_report_title(row)
        return _link(link, title) if link and title else html.escape(title or "")
    if value in (None, "", []):
        return ""
    if field == "company_name":
        return html.escape(_display_company_name(value, row))
    if field in {"core_basis", "top_position_weight_changes"}:
        return _format_bullets(str(value))
    if field in {"source_url", "report_link"}:
        return _link(str(value), "열기")
    if field == "date":
        return html.escape(_short_date(value))
    if field == "generated_at_kst":
        return html.escape(_short_datetime(value))
    if isinstance(value, list):
        return "".join(f"<span class='tag'>{html.escape(str(item))}</span>" for item in value if item not in (None, ""))
    if field in {"relative_volume", "avg_relative_volume"}:
        number = _number(value)
        return "" if number is None else html.escape(f"{int(round(number))}x")
    if field in PERCENT_FIELDS:
        number = _number(value)
        return "" if number is None else html.escape(f"{int(round(number))}%")
    if field in NUMERIC_FIELDS or isinstance(value, (int, float)):
        return html.escape(_compact_number(value))
    return html.escape(_clean_display_text(str(value)))


def _cell_title(field: str, value: Any, row: dict[str, Any]) -> str:
    if field == "recent_report_title":
        return _display_report_title(row) or ""
    if field == "recent_report_broker":
        return _display_report_broker(row) or ""
    if field == "company_name" and value not in (None, "", []):
        return _display_company_name(value, row)
    if field in {"core_basis", "top_position_weight_changes"} and value not in (None, "", []):
        return " / ".join(_basis_parts(str(value)))
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item not in (None, ""))
    if field in {"future_industry_theme", "signals"} and value not in (None, "", []):
        return _clean_display_text(str(value))
    return ""


def _display_company_name(value: Any, row: dict[str, Any]) -> str:
    text = _clean_display_text(str(value))
    if row.get("country_code") == "KR" or row.get("country") == "한국":
        return koreanize_kr_company_name(text) or text
    return text


def _link(url: str, label: str) -> str:
    return f"<a href='{html.escape(url, quote=True)}' target='_blank' rel='noopener'>{html.escape(label)}</a>"


def _format_bullets(value: str) -> str:
    parts = _basis_parts(value)
    if len(parts) <= 1:
        return _highlight_keywords(_clean_display_text(value))
    items = "".join(f"<span class='{_bullet_class(part)}'>{_highlight_keywords(part)}</span>" for part in parts)
    return f"<div class='bullets'>{items}</div>"


def _bullet_class(value: str) -> str:
    text = str(value)
    if re.search(r"리스크|감소|청산|하락|부채|공매도|-\d", text):
        return "b-risk"
    if re.search(r"외국인|기관|순매수|수급|신규|증가|보유", text):
        return "b-flow"
    if re.search(r"거래량|신고가|52주|급증|돌파", text):
        return "b-hot"
    if re.search(r"RSI|ADX|베타|갭|이동평균|성과", text, re.IGNORECASE):
        return "b-tech"
    if re.search(r"FCF|ROIC|ROE|성장|목표가|상승|마진|저평가", text, re.IGNORECASE):
        return "b-good"
    return "b-neutral"


def _highlight_keywords(value: str) -> str:
    text = _clean_display_text(str(value))
    pattern = re.compile(
        r"(거래량 급증|52주 신고가|상대거래량|거래량|신고가|52주|외국인|기관|순매수|수급|신규|증가|보유|FCF마진|FCF|ROIC|ROE|성장|목표가|상승|마진|성과|RSI|ADX|베타|갭|리스크|감소|청산|하락|부채비율|부채|공매도|-\d+%)",
        re.IGNORECASE,
    )
    output: list[str] = []
    last = 0
    for match in pattern.finditer(text):
        output.append(html.escape(text[last : match.start()]))
        token = match.group(0)
        output.append(f"<strong class='kw {_keyword_class(token)}'>{html.escape(token)}</strong>")
        last = match.end()
    output.append(html.escape(text[last:]))
    return "".join(output)


def _keyword_class(value: str) -> str:
    text = value.lower()
    if re.search(r"리스크|감소|청산|하락|부채|공매도|-", text):
        return "kw-risk"
    if re.search(r"외국인|기관|순매수|수급|신규|증가|보유", text):
        return "kw-flow"
    if re.search(r"거래량|신고가|52주|급증", text):
        return "kw-hot"
    if re.search(r"rsi|adx|베타|갭|성과", text):
        return "kw-tech"
    return "kw-good"


def _compact_number(value: Any) -> str:
    number = _number(value)
    if number is None:
        return ""
    sign = "-" if number < 0 else ""
    absolute = abs(number)
    for threshold, suffix in (
        (1_000_000_000_000, "T"),
        (1_000_000_000, "B"),
        (1_000_000, "M"),
        (1_000, "K"),
    ):
        if absolute >= threshold:
            return f"{sign}{int(round(absolute / threshold)):,}{suffix}"
    rounded = int(round(absolute))
    if rounded == 0:
        return "0"
    return f"{sign}{rounded:,}"


def _clean_display_text(value: str) -> str:
    def repl(match: Any) -> str:
        number = float(match.group(0).replace(",", ""))
        return f"{int(round(number)):,}"

    return re.sub(r"-?\d[\d,]*\.\d+", repl, value)


def _short_date(value: Any) -> str:
    text = str(value or "").strip()
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", text)
    if not match:
        return text
    return f"{int(match.group(2))}/{int(match.group(3))}"


def _short_datetime(value: Any) -> str:
    text = str(value or "").strip()
    match = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})(\s+\d{2}:\d{2}:\d{2})", text)
    if not match:
        return _short_date(text)
    return f"{int(match.group(2))}/{int(match.group(3))}{match.group(4)}"


def _display_report_broker(row: dict[str, Any]) -> str | None:
    broker = row.get("recent_report_broker")
    if broker not in (None, "", []) and not _has_generic_report(row):
        return _clean_display_text(str(broker))
    return _fallback_report_broker(row)


def _display_report_title(row: dict[str, Any]) -> str | None:
    title = row.get("recent_report_title")
    if title not in (None, "", []) and not _is_generic_report_title(title):
        return _clean_display_text(str(title))
    return _fallback_report_title(row)


def _has_generic_report(row: dict[str, Any]) -> bool:
    return _is_generic_report_title(row.get("recent_report_title")) or _is_generic_report_link(row.get("report_link"))


def _is_generic_report_title(value: Any) -> bool:
    text = re.sub(r"\s+", " ", str(value or "")).strip().lower()
    return text in {"컨센서스 검색", "analyst estimates / analysis", "analyst estimates/analysis"}


def _is_generic_report_link(value: Any) -> bool:
    text = str(value or "")
    return "markets.hankyung.com/consensus?searchWord=" in text or re.search(r"finance\.yahoo\.com/quote/[^/]+/analysis/?", text) is not None


def _fallback_report_broker(row: dict[str, Any]) -> str | None:
    if row.get("country_code") == "US" or row.get("country") == "미국":
        return "Yahoo Finance News"
    if row.get("country_code") == "KR" or row.get("country") == "한국":
        return "네이버증권 뉴스"
    return None


def _fallback_report_title(row: dict[str, Any]) -> str | None:
    if row.get("country_code") == "US" or row.get("country") == "미국":
        return "Yahoo News"
    if row.get("country_code") == "KR" or row.get("country") == "한국":
        return "종목 뉴스"
    return None


def _fallback_report_link(row: dict[str, Any]) -> str | None:
    if row.get("report_link") and not _is_generic_report_link(row.get("report_link")):
        return str(row["report_link"])
    ticker = row.get("ticker")
    if not ticker:
        return None
    if row.get("country_code") == "US" or row.get("country") == "미국":
        return f"https://finance.yahoo.com/quote/{ticker}/news/"
    if row.get("country_code") == "KR" or row.get("country") == "한국":
        code = _report_ticker_code(ticker)
        return f"https://finance.naver.com/item/news.naver?code={code}" if code else None
    return None


def _report_ticker_code(value: Any) -> str | None:
    match = re.search(r"\d{6}", str(value or ""))
    return match.group(0) if match else None


def _report_company_query(row: dict[str, Any]) -> str | None:
    for key in ("company_name", "naver_title", "ticker"):
        value = row.get(key)
        if value in (None, ""):
            continue
        text = re.sub(r"\s*:\s*(?:Npay\s*)?증권.*$", "", str(value)).strip()
        text = re.sub(r"\s*:\s*네이버.*$", "", text).strip()
        if text and "\ufffd" not in text:
            return text
    return None


def _number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def _style_tokens(field: str, value: Any) -> list[str]:
    tokens: list[str] = []
    number = _number(value)

    if field in {"ticker", "cusip"}:
        tokens.append("key-cell")
    if field == "company_name":
        tokens.append("company-cell")
    if field == "section":
        tokens.append("section-cell")
    if field == "country":
        label = str(value or "")
        if "미국" in label or label.upper() == "US":
            tokens.append("country-us")
        elif "한국" in label or label.upper() == "KR":
            tokens.append("country-kr")
    if field == "signals":
        tokens.append("signal-cell")
    if field == "future_industry_theme":
        tokens.append("theme-cell")
    if field == "core_basis":
        tokens.append("basis-cell")
    if field in {"recent_report_title", "report_link", "source_url"}:
        tokens.append("report-cell")
    if field in {
        "investment_priority_score",
        "long_future_score",
        "leading_supply_score",
        "institution_flow_score",
        "foreign_flow_investment_score",
        "famous_13f_score",
        "avg_investment_priority_score",
        "avg_long_future_score",
        "avg_leading_supply_score",
    }:
        tokens.append("score-cell")
    if field in {"forward_per", "forward_peg", "trailing_per", "pbr", "price_to_sales", "price_to_fcf", "enterprise_to_revenue", "enterprise_to_ebitda"}:
        tokens.append("valuation-cell")
    if field in {"revenue_growth_yoy", "revenue_growth_qoq", "eps_growth_yoy", "eps_growth_qoq", "expected_revenue_growth", "expected_eps_growth"}:
        tokens.append("growth-cell")
    if field in {"fcf_margin", "roic", "roe", "roa", "gross_margin", "operating_margin", "profit_margin", "current_ratio", "quick_ratio"}:
        tokens.append("quality-cell")
    if field in {"distance_to_52w_high_pct", "position_52w_pct", "sma50_gap_pct", "sma200_gap_pct", "change_pct", "change_from_open_pct", "gap_pct", "performance_1w", "performance_1m", "performance_3m", "performance_6m", "performance_ytd", "performance_1y", "rsi_14", "adx_14", "atr_14", "beta"}:
        tokens.append("momentum-cell")
    if field in {"risk_penalty", "debt_to_equity", "short_percent_float", "volatility_d", "decreased_institution_count", "exited_institution_count"}:
        tokens.append("risk-cell")
    if field in {"market_cap", "enterprise_value", "total_revenue", "total_cash", "total_debt"}:
        tokens.append("size-cell")
    if field in {"volume", "average_volume_30d", "relative_volume", "avg_relative_volume", "float_shares", "shares_outstanding", "foreign_rank_volume", "institution_rank_volume"}:
        tokens.append("liquidity-cell")
    if field in {"foreign_ownership_rate", "institutional_ownership_pct", "insider_ownership_pct", "foreign_ownership_change_20d"}:
        tokens.append("ownership-cell")

    score_fields = {
        "investment_priority_score",
        "long_future_score",
        "leading_supply_score",
        "institution_flow_score",
        "foreign_flow_investment_score",
        "famous_13f_score",
        "avg_investment_priority_score",
        "avg_long_future_score",
        "avg_leading_supply_score",
    }
    if field in score_fields and number is not None:
        if number >= 8:
            tokens.append("pos-strong")
        elif number >= 5:
            tokens.append("pos-buy")

    if field in {"relative_volume", "avg_relative_volume"} and number is not None:
        if number >= 5:
            tokens.append("volume-hot")
        elif number >= 3:
            tokens.append("volume-watch")
        elif number >= 2:
            tokens.append("warn")
    if field == "market_cap" and number is not None:
        if number >= 1_000_000_000_000:
            tokens.append("metric-purple")
        elif number >= 10_000_000_000:
            tokens.append("metric-blue")
        elif number > 0:
            tokens.append("metric-soft")
    if field in {"volume", "average_volume_30d", "float_shares", "shares_outstanding", "foreign_rank_volume", "institution_rank_volume"} and number is not None and number > 0:
        tokens.append("metric-soft")
    if field == "analyst_opinion" and value not in (None, "", []):
        text = str(value)
        if "매수" in text or "Buy" in text:
            tokens.append("pos-buy")
        else:
            tokens.append("metric-soft")

    if field == "distance_to_52w_high_pct" and number is not None:
        if number >= -1:
            tokens.append("pos-strong")
        elif number >= -3:
            tokens.append("pos-buy")
        elif number <= -10:
            tokens.append("warn")
    if field == "position_52w_pct" and number is not None:
        if number >= 90:
            tokens.append("pos-strong")
        elif number >= 70:
            tokens.append("pos-buy")
        elif number <= 25:
            tokens.append("warn")

    if field in {"sma50_gap_pct", "sma200_gap_pct", "change_pct", "change_from_open_pct", "gap_pct"} and number is not None:
        if number >= 10:
            tokens.append("volume-hot")
        elif number > 0:
            tokens.append("pos-buy")
        elif number <= -10:
            tokens.append("neg")

    if field in {"performance_1w", "performance_1m", "performance_3m", "performance_6m", "performance_ytd", "performance_1y"} and number is not None:
        if number >= 25:
            tokens.append("pos-strong")
        elif number >= 10:
            tokens.append("pos-buy")
        elif number > 0:
            tokens.append("metric-soft")
        elif number < 0:
            tokens.append("neg")
    if field == "volatility_d" and number is not None:
        if number >= 12:
            tokens.append("neg")
        elif number >= 5:
            tokens.append("warn")
    if field == "rsi_14" and number is not None:
        if number >= 75:
            tokens.append("warn")
        elif number >= 55:
            tokens.append("pos-buy")
        elif number <= 30:
            tokens.append("metric-cyan")
    if field == "adx_14" and number is not None:
        if number >= 35:
            tokens.append("pos-strong")
        elif number >= 25:
            tokens.append("pos-buy")
    if field == "atr_14" and number is not None and number > 0:
        tokens.append("metric-soft")

    if field in {"target_upside_pct", "average_change_pct", "foreign_ownership_change_20d"} and number is not None:
        if number >= 20:
            tokens.append("pos-strong")
        elif number > 0:
            tokens.append("pos-buy")
        elif number < 0:
            tokens.append("neg")

    flow_fields = {
        "foreign_net_buy",
        "foreign_net_buy_5d",
        "foreign_net_buy_20d",
        "foreign_net_buy_amount_mil_krw",
        "institution_net_buy",
        "institution_net_buy_5d",
        "institution_net_buy_20d",
        "institution_net_buy_amount_mil_krw",
        "net_supply_5d",
        "net_supply_20d",
        "total_share_change",
        "total_current_shares",
    }
    if field in flow_fields and number is not None:
        if number > 0:
            tokens.append("flow-pos")
        elif number < 0:
            tokens.append("neg")

    if field in {"new_institution_count", "increased_institution_count", "position_weight_increased_count"} and number is not None and number > 0:
        tokens.append("pos-buy")
    if field in {"guru_position_score", "max_position_weight_change_pct"} and number is not None and number > 0:
        tokens.append("pos-strong")
    if field == "average_position_weight_change_pct" and number is not None:
        if number > 0:
            tokens.append("pos-buy")
        elif number < 0:
            tokens.append("neg")
    if field == "decreased_institution_count" and number is not None and number > 0:
        tokens.append("warn")
    if field == "exited_institution_count" and number is not None and number > 0:
        tokens.append("neg")

    if field in {"forward_per", "trailing_per", "price_to_sales", "price_to_fcf"} and number is not None:
        if 0 < number <= 15:
            tokens.append("metric-blue")
        elif 0 < number <= 25:
            tokens.append("metric-soft")
        elif number >= 60:
            tokens.append("neg")
        elif number >= 40:
            tokens.append("warn")
    if field == "forward_peg" and number is not None:
        if 0 < number <= 1:
            tokens.append("metric-purple")
        elif 0 < number <= 2:
            tokens.append("pos-buy")
        elif number >= 4:
            tokens.append("neg")
        elif number > 2:
            tokens.append("warn")
    if field == "pbr" and number is not None:
        if 0 < number <= 1:
            tokens.append("metric-blue")
        elif 0 < number <= 2:
            tokens.append("metric-soft")
        elif number >= 8:
            tokens.append("neg")
        elif number >= 4:
            tokens.append("warn")
    if field == "beta" and number is not None:
        if number >= 2:
            tokens.append("warn")
        elif 0 < number <= 1:
            tokens.append("metric-cyan")

    growth_fields = {
        "revenue_growth_yoy",
        "revenue_growth_qoq",
        "eps_growth_yoy",
        "eps_growth_qoq",
        "expected_revenue_growth",
        "expected_eps_growth",
    }
    if field in growth_fields and number is not None:
        if number >= 25:
            tokens.append("pos-strong")
        elif number >= 10:
            tokens.append("pos-buy")
        elif number > 0:
            tokens.append("metric-soft")
        elif number < 0:
            tokens.append("neg")

    if field in {"fcf_margin", "roic", "roe", "roa", "gross_margin", "operating_margin", "profit_margin"} and number is not None:
        if number >= 15:
            tokens.append("pos-strong")
        elif number >= 8:
            tokens.append("metric-cyan")
        elif number > 0:
            tokens.append("metric-soft")
        elif number < 0:
            tokens.append("neg")
    if field == "debt_to_equity" and number is not None:
        if number >= 200:
            tokens.append("neg")
        elif number >= 100:
            tokens.append("warn")
        elif number >= 0:
            tokens.append("metric-cyan")
    if field in {"current_ratio", "quick_ratio"} and number is not None:
        if number >= 2:
            tokens.append("pos-buy")
        elif number < 1:
            tokens.append("warn")
    if field == "risk_penalty" and number is not None:
        if number >= 2:
            tokens.append("neg")
        elif number >= 1:
            tokens.append("warn")
        elif number <= 0.5:
            tokens.append("pos-buy")
    if field in {"dividend_yield", "foreign_ownership_rate", "institutional_ownership_pct", "insider_ownership_pct"} and number is not None and number > 0:
        tokens.append("metric-soft")
    if field == "short_percent_float" and number is not None:
        if number >= 20:
            tokens.append("neg")
        elif number >= 10:
            tokens.append("warn")

    return tokens


def _classes(field: str, value: Any) -> str:
    classes = []
    if field in NUMERIC_FIELDS or field in PERCENT_FIELDS or field in {"relative_volume", "avg_relative_volume"}:
        classes.append("num")
    if field == "core_basis":
        classes.append("basis")
    classes.extend(_style_tokens(field, value))
    return " ".join(dict.fromkeys(classes))


def _merge_daily_tracking(path: Path, summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing: list[dict[str, Any]] = []
    if path.exists():
        try:
            existing = pd.read_csv(path).to_dict("records")
        except Exception:
            existing = []
    combined = [_normalize_tracking_row(row) for row in summary_rows + existing]
    seen = set()
    unique = []
    for row in combined:
        key = (str(row.get("date")), str(row.get("section")))
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    unique.sort(key=lambda row: str(row.get("date") or ""), reverse=True)
    return unique[:500]


def _normalize_tracking_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    section = str(normalized.get("section") or "")
    normalized["section"] = LEGACY_SECTION_TITLES.get(section, section)
    return normalized


def _dedupe_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique = []
    seen = set()
    for row in rows:
        key = row.get("ticker") or row.get("cusip") or row.get("company_name") or row.get("future_industry_theme") or tuple(sorted(row.items()))
        key = str(key).upper()
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _write_xlsx(path: Path, sections: dict[str, list[dict[str, Any]]]) -> None:
    ensure_dir(path.parent)
    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for key, title in SECTION_TITLES:
                raw_rows = sections.get(key, [])
                columns = _columns_for_rows(SECTION_COLUMNS.get(key, STOCK_COLUMNS), raw_rows)
                rows = [_export_row(row, columns) for row in raw_rows]
                frame = pd.DataFrame(rows, columns=[label for _, label in columns])
                frame.to_excel(writer, sheet_name=title[:31], index=False)
                _style_xlsx_sheet(writer.sheets[title[:31]], columns, raw_rows)
    except Exception:
        return


def _style_xlsx_sheet(sheet: Any, columns: list[tuple[str, str]], raw_rows: list[dict[str, Any]]) -> None:
    border = Border(bottom=Side(style="thin", color="D8DFE9"))
    header_fill = PatternFill("solid", fgColor="2F2F2F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, (field, label) in enumerate(columns, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = _xlsx_width(field, label)
        for row_index, raw in enumerate(raw_rows, start=2):
            cell = sheet.cell(row=row_index, column=column_index)
            tokens = _style_tokens(field, raw.get(field))
            _apply_xlsx_cell_style(cell, tokens, field)
            if field == "recent_report_title":
                link = _fallback_report_link(raw)
                if link:
                    cell.hyperlink = link
                    cell.font = Font(color="0563C1", bold=True, underline="single")
            if field in NUMERIC_FIELDS or field in PERCENT_FIELDS or field in {"relative_volume", "avg_relative_volume"}:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border


def _xlsx_width(field: str, label: str) -> int:
    widths = {
        "date": 12,
        "country": 10,
        "ticker": 13,
        "company_name": 18,
        "signals": 20,
        "core_basis": 34,
        "future_industry_theme": 20,
        "top_tickers": 34,
        "stock_count": 12,
        "avg_investment_priority_score": 16,
        "avg_long_future_score": 16,
        "avg_leading_supply_score": 16,
        "avg_relative_volume": 14,
        "total_revenue": 16,
        "free_cash_flow": 16,
        "foreign_net_buy_5d": 16,
        "foreign_net_buy_20d": 16,
        "foreign_net_buy": 16,
        "foreign_net_buy_amount_mil_krw": 18,
        "foreign_rank_volume": 16,
        "institution_net_buy_5d": 16,
        "institution_net_buy_20d": 16,
        "institution_net_buy": 16,
        "institution_net_buy_amount_mil_krw": 18,
        "institution_rank_volume": 16,
        "supply_source": 22,
        "net_supply_5d": 16,
        "net_supply_20d": 16,
        "float_shares": 16,
        "shares_outstanding": 16,
        "institutional_ownership_pct": 14,
        "insider_ownership_pct": 14,
        "short_percent_float": 14,
        "recent_report_title": 28,
        "recent_report_broker": 18,
        "institutions": 28,
        "section": 22,
        "generated_at_kst": 20,
    }
    return widths.get(field, min(max(len(label) + 4, 12), 20))


def _apply_xlsx_cell_style(cell: Any, tokens: list[str], field: str) -> None:
    style = _xlsx_style(tokens)
    if style:
        fill, font_color, bold = style
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=font_color, bold=bold)
    elif field in {"ticker", "company_name", "future_industry_theme", "core_basis", "top_tickers"}:
        cell.font = Font(color="111827", bold=True)


def _xlsx_style(tokens: list[str]) -> tuple[str, str, bool] | None:
    palette = {
        "pos-strong": ("15803D", "FFFFFF", True),
        "pos-buy": ("DCFCE7", "166534", True),
        "warn": ("FEE2E2", "991B1B", True),
        "neg": ("FCA5A5", "7F1D1D", True),
        "volume-hot": ("15803D", "FFFFFF", True),
        "volume-watch": ("DCFCE7", "166534", True),
        "flow-pos": ("16A34A", "FFFFFF", True),
        "score-cell": ("F8FAFC", "334155", True),
        "valuation-cell": ("F8FAFC", "334155", True),
        "growth-cell": ("F8FAFC", "334155", True),
        "quality-cell": ("F8FAFC", "334155", True),
        "momentum-cell": ("F8FAFC", "334155", True),
        "risk-cell": ("FEF2F2", "991B1B", True),
        "liquidity-cell": ("F8FAFC", "334155", True),
        "size-cell": ("F8FAFC", "334155", True),
        "ownership-cell": ("F8FAFC", "334155", True),
        "metric-blue": ("DCFCE7", "166534", True),
        "metric-cyan": ("DCFCE7", "166534", True),
        "metric-purple": ("DCFCE7", "166534", True),
        "metric-soft": ("F0FDF4", "166534", True),
        "key-cell": ("F0FDF4", "14532D", True),
        "company-cell": ("FFFFFF", "111827", True),
        "country-us": ("E0F2FE", "075985", True),
        "country-kr": ("F0FDF4", "166534", True),
        "signal-cell": ("F8FAFC", "475569", True),
        "theme-cell": ("EDE9FE", "5B21B6", True),
        "basis-cell": ("F8FAFC", "334155", True),
        "report-cell": ("EFF6FF", "0563C1", True),
        "section-cell": ("ECFDF5", "047857", True),
    }
    priority = [
        "neg",
        "warn",
        "pos-strong",
        "volume-hot",
        "pos-buy",
        "volume-watch",
        "flow-pos",
        "score-cell",
        "risk-cell",
        "liquidity-cell",
        "growth-cell",
        "quality-cell",
        "momentum-cell",
        "valuation-cell",
        "size-cell",
        "ownership-cell",
        "metric-purple",
        "metric-blue",
        "metric-cyan",
        "metric-soft",
        "key-cell",
        "country-us",
        "country-kr",
        "theme-cell",
        "report-cell",
        "section-cell",
        "basis-cell",
        "signal-cell",
        "company-cell",
    ]
    for token in priority:
        if token in tokens:
            return palette[token]
    return None


def _export_row(row: dict[str, Any], columns: list[tuple[str, str]]) -> dict[str, str]:
    return {label: _format_export_cell(field, row.get(field), row) for field, label in columns}


def _format_export_cell(field: str, value: Any, row: dict[str, Any]) -> str:
    if field == "recent_report_broker":
        return _display_report_broker(row) or ""
    if field == "recent_report_title":
        return _display_report_title(row) or ""
    if value in (None, "", []):
        return ""
    if field == "date":
        return _short_date(value)
    if field == "generated_at_kst":
        return _short_datetime(value)
    if field == "company_name":
        return _display_company_name(value, row)
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item not in (None, ""))
    if field in {"core_basis", "top_position_weight_changes"}:
        return "\n".join(f"- {part}" for part in _basis_parts(str(value)))
    if field in {"relative_volume", "avg_relative_volume"}:
        number = _number(value)
        return "" if number is None else f"{int(round(number))}x"
    if field in PERCENT_FIELDS:
        number = _number(value)
        return "" if number is None else f"{int(round(number))}%"
    if field in NUMERIC_FIELDS or isinstance(value, (int, float)):
        return _compact_number(value)
    return _clean_display_text(str(value))


def _basis_parts(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s+/\s+", _clean_display_text(value)) if part.strip()]
